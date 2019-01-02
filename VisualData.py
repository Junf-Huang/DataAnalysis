import openpyxl

from wordcloud import WordCloud
from matplotlib import pyplot as plt 

import plotly 
import plotly.plotly as py
import plotly.graph_objs as go
plotly.tools.set_credentials_file(username='Junf-Huang', api_key='UcLG1VBERu4MiPHw7wXp')


wb = openpyxl.load_workbook('gzjs/phone.xlsx')
sheet = wb["statistic"]

# 读取第一行有关统计信息
start_row = 1
name = sheet.cell(row=start_row, column=1).value
salesVolume = sheet.cell(row=start_row, column=2).value
print('name:', name)
print('salesVolume:', salesVolume)

# 从第二行开始，记录统计数据  
# 数据过多，截取前20
x = []
y = []
start_row += 1
end_row = 20
print("row:", sheet.max_row)
print("col:", sheet.max_column)
# range最后一个范围得加 1
for i in range(start_row, end_row + 1):
    x.append(sheet.cell(row=i, column=1).value)
    y.append(sheet.cell(row=i, column=2).value)

print ("x: ", x)
print ("y: ", y)


# 读取品牌、销量的条形图、饼图
# python wordcloud 
data = [go.Bar(
            x = x,
            y = y,
            name = salesVolume,
            width = [0.8],
            marker = dict(color='rgb(31, 119, 180)')
    )]

py.iplot(data, filename='basic-bar')


trace = go.Pie(labels=x, values=y)
py.iplot([trace], filename='basic_pie_chart')

sheet = wb["comment"]
#情感倾向图
title = []
values = []
strings = []
label = ['好', '坏']
start_row += 1
end_row = 4
for i in range(start_row, end_row + 1):
    title.append(sheet.cell(row=i, column=1).value)
    values.append(float(sheet.cell(row=i, column=2).value))
    strings.append(sheet.cell(row=i, column=3).value)

fig = {
    'data': [
        {
            'labels': label,
            'values': [values[0], 1-values[0]],
            'type': 'pie',
            'name': title[0],
            'marker': {'colors': ['rgb(255, 127, 14)',
                                  'rgb(31, 119, 180)']},
            'domain': {'x': [0, .48],
                       'y': [0, .49]},
            'hoverinfo':'label+percent+name',
            'textinfo':'none'
        },
        {
            'labels': label,
            'values': [values[1], 1-values[1]],
            'marker': {'colors': ['rgb(255, 127, 14)',
                                  'rgb(31, 119, 180)']},
            'type': 'pie',
            'name':title[1],
            'domain': {'x': [.52, 1],
                       'y': [0, .49]},
            'hoverinfo':'label+percent+name',
            'textinfo':'none'
        }
    ],
    'layout': {'title': 'Commodity Sentiment Analysis',
               'showlegend': False}
}

py.iplot(fig, filename='pie_chart_subplots')

#典型意见词云图
for string in strings:  
    font = '/usr/share/fonts/adobe-source-han-sans/SourceHanSansCN-Regular.otf'
    wc = WordCloud(font_path=font, #如果是中文必须要添加这个，否则会显示成框框
                   background_color='white',
                   width=1000,
                   height=800,
                   ).generate(string)
    wc.to_file('word.png') #保存图片
    plt.imshow(wc)  #用plt显示图片
    plt.axis('off') #不显示坐标轴
    plt.show() #显示图片
