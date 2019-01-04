import openpyxl

import pandas as pd
import numpy as np

import plotly 
import plotly.plotly as py
import plotly.graph_objs as go
plotly.tools.set_credentials_file(username='Junf-Huang', api_key='UcLG1VBERu4MiPHw7wXp')


wb = openpyxl.load_workbook('gzjs/phone.xlsx')
sheet = wb["phoneDetail"]

# 读取第一行有关统计信息
start_row = 1
price = sheet.cell(row=start_row, column=3).value
salesVolume = sheet.cell(row=start_row, column=4).value
print('price:', price)
print('salesVolume:', salesVolume)

start_row += 1
x = []
y = []

for i in range(start_row, sheet.max_row + 1):
    x.append(int(float(sheet.cell(row=i, column=3).value)))
    y.append(int(sheet.cell(row=i, column=4).value))

x = np.array(x) 
y = np.array(y) 
df = pd.DataFrame({price:x, salesVolume:y})
print(df)
# 用来衡量两个数据集合是否在一条线上面，用来衡量定距变量间的线性关系
# 相关系数越接近于1或-1，相关度越强，相关系数越接近于0，相关度越弱。相关系数 0.8-1.0 极强相关
print('pearson:\n', df.corr())
# 前提>0.467,系数为正 ，两个属性正相关,系数为负 ，两个属性负相关
print('kendall:\n', df.corr('kendall'))

# Create a trace
trace = go.Scatter(
    x = x,
    y = y,
    mode = 'markers'
)

data = [trace]

# Plot and embed in ipython notebook!
py.iplot(data, filename='basic-scatter')

