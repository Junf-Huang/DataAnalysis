# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import openpyxl
import re
import pandas as pd

class GzjsPipeline(object):
    def process_item(self, item, spider):
        line = [item['pid'], item['name'], item['price'], item['salesVolume']]
        if(float(line[2]) >= 0):
            # 保存全部商品信息
            self.sheet = self.wb['phoneDetail']
            self.sheet.append(line)
            self.wb.save('phone.xlsx')
        
            # 统计总销量
            key = item['brand']
            number = item['salesVolume']
            if key not in self.dict.keys():
                self.dict[key] = number
            else:
                self.dict[key] += number
            
        return item

    def open_spider(self, spider):
        self.wb = openpyxl.Workbook()
        #获取活跃的sheet
        self.sheet = self.wb.active
        # 直接赋值就可以改工作表的名称
        self.sheet.title = "phoneDetail"
        self.sheet.append(['商品编号', '商品名称', '价格', '销量'])

        # 初始化统计
        self.dict = {}
        self.sheet = self.wb.create_sheet('statistic', 0)
        self.sheet.append(['品牌', '总销量'])

    def close_spider(self, spider):
        print("####记录统计数据####")
        # 选择sheet表
        self.sheet = self.wb['statistic']
        # 排序销量
        # 如果想排序excel数据,用pandas
        self.dict = sorted(self.dict.items(),key = lambda x:x[1],reverse = True)
        # print("销量统计排序：", self.dict)
        for i in self.dict:
            list = []
            for j in i:
                list.append(j)
            self.sheet.append(list)
        self.wb.save('phone.xlsx')


class CommentPipeline(object):
    def process_item(self, item, spider):
        line = [item['pid'], item['sentiment'], item['keywords']]
        self.sheet = self.wb['comment']
        self.sheet.append(line)
        self.wb.save('phone.xlsx')
        return item

    def open_spider(self, spider):
        self.wb = openpyxl.load_workbook('phone.xlsx')
        self.sheet = self.wb.create_sheet('comment', 0)
        self.sheet.append(['手机编号', '好评率','典型意见'])
        self.wb.save('phone.xlsx')

    def close_spider(self, spider):
        pass